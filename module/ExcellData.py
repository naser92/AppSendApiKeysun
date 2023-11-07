import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter
import shutil
import os
import sys
sys.path.append(os.getcwd())
from model.invoiceModel import InvoiceData
from model.columns import NameColumnsInvoic,Columns
col = NameColumnsInvoic()
import uuid
from model.setting import VersionApp
import jdatetime
import concurrent.futures

class ExcellData ():
    def __init__(self, path) -> None:
        self.path = path
          
    
    def getInvoice(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.sheetNames[0], dtype=str)
            data = data.replace(np.nan, '')
            invoice = np.array(data)
            return invoice
        except:
            return []

    def getInvoiceItem(self):
        try:
            data = pd.read_excel(self.path,sheet_name= self.sheetNames[1], dtype=str)
            data = data.replace(np.nan, '')
            item = np.array(data)
            return item
        except:
            return []
    
    def getPayment(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.sheetNames[2], dtype=str)
            data = data.replace(np.nan, '')
            pay = np.array(data)
            return pay
        except:
            return []
        
    def getRevokeInvoice(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.name_column_revoke, dtype=str)
            data = data.replace(np.nan, '')
            pay = np.array(data)
            return pay
        except:
            return []
        
    def getDeleteUniquId(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.name_column_delete, dtype=str)
            data = data.replace(np.nan, '')
            pay = np.array(data)
            return pay
        except:
            return []
        
        
    def getBillInvoice(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.name_column_Bill, dtype=str)
            data = data.replace(np.nan, '')
            pay = np.array(data)
            return pay
        except:
            return []
    
        
    def checkExcel(self, paternType:int):
        try:
            data = openpyxl.load_workbook(self.path)
            result = []
            if len(data.sheetnames) > 3 or len(data.sheetnames) < 1:
                return result
            
            index = 0
            for sheet in data.sheetnames:
                columns = data.get_sheet_by_name(sheet)
                if index == 0 and paternType == 1:
                    if columns.max_column == 21 : 
                        result.append(1) 
                        self.name_column_invoice = sheet
                        sh = data.active
                        self.invoiceFactorNumber = sh['A1'].value
                    else: result.append(0) 
                elif index == 0 and paternType == 2:
                    if columns.max_column == 18 : 
                        result.append(1) 
                        self.name_column_invoice = sheet
                        sh = data.active
                        self.invoiceFactorNumber = sh['A1'].value
                    else: result.append(0) 
                elif index == 1 :
                    if columns.max_column == 21 : 
                        result.append(1) 
                        self.name_column_invoiceItem = sheet
                        sh = data.active
                        self.invoiceItemFactorNumber = sh['A1'].value
                    else: result.append(0) 
                elif index == 2 and paternType == 1 : 
                    if columns.max_column == 11 : 
                        result.append(1) 
                        self.name_column_payment = sheet
                        sh = data.active
                        self.PaymentFactorNumber = sh['A1'].value
                    else: result.append(0) 
                
                index += 1
            return result
        except:
            return None
        
    def excelCheckRevok(self):
        try:
            data = openpyxl.load_workbook(self.path)
            flag = False
            if len(data.sheetnames) > 1:
                return flag
            sheet = data.sheetnames[0]
            columns = data.get_sheet_by_name(sheet)
            if columns.max_column == 4:
                self.name_column_revoke = sheet
                flag = True
            
            return flag

        except:
            return False
        
    def excelCheckBill(self):
        try:
            data = openpyxl.load_workbook(self.path)
            flag = False
            if len(data.sheetnames) < 1:
                return flag
            sheet = data.sheetnames[0]
            columns = data.get_sheet_by_name(sheet)
            if columns.max_column == 16:
                self.name_column_Bill = sheet
                flag = True
            
            return flag

        except:
            return False
        
    def excelCheckDelete(self):
        try:
            data = openpyxl.load_workbook(self.path)
            flag = False
            sheet = data.sheetnames[0]
            columns = data.get_sheet_by_name(sheet)
            if columns.max_column == 1:
                self.name_column_delete = sheet
                flag = True
            return flag
        except:
            return False
    
    @classmethod
    def getLetter(col):
        return get_column_letter(col)
    
    def copyFile(self,path):
        try:
            new  = path[:-5] + "_read.xlsx"
            result = path[:-5] + "_result.xlsx"
            shutil.copyfile(path, new,follow_symlinks=True)
            shutil.copyfile(path, result,follow_symlinks=True)
            os.remove(path)
            self.path = result
        except:
            self.path = path
    
    def preparationExcellN11(self):
        self.copyFile(self.path)
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(22)+'1'] = "uniqueId"
        worksheet[get_column_letter(23)+'1'] = "Status Request Server"
        worksheet[get_column_letter(24)+'1'] = "taxSerialNumber"
        worksheet[get_column_letter(25)+'1'] = "message"
        worksheet[get_column_letter(26)+'1'] = "FieldError"
        workbook.save(self.path)



    def preparationExcellN21(self):
        self.copyFile(self.path)
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(17)+'1'] = "uniqueId"
        worksheet[get_column_letter(18)+'1'] = "Status Request Server"
        worksheet[get_column_letter(19)+'1'] = "taxSerialNumber"
        worksheet[get_column_letter(20)+'1'] = "message"
        worksheet[get_column_letter(21)+'1'] = "FieldError"
        workbook.save(self.path)
        
    def SaveResultN11(self,data,index):
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(22)+str(index+1)] = data[0]#"uniqueId"
        worksheet[get_column_letter(23)+str(index+1)] = data[1]#"Status Request Server"
        worksheet[get_column_letter(24)+str(index+1)] = data[2]#"taxSerialNumber"
        worksheet[get_column_letter(25)+str(index+1)] = data[3]#"message"
        worksheet[get_column_letter(26)+str(index+1)] = data[4]#"FieldError"
        workbook.save(self.path)

    def SaveResultN21(self,data,index):
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(17)+str(index+1)] = data[0]#"uniqueId"
        worksheet[get_column_letter(18)+str(index+1)] = data[1]#"Status Request Server"
        worksheet[get_column_letter(19)+str(index+1)] = data[2]#"taxSerialNumber"
        worksheet[get_column_letter(20)+str(index+1)] = data[3]#"message"
        worksheet[get_column_letter(21)+str(index+1)] = data[4]#"FieldError
        workbook.save(self.path)
    

    def data_generator(self, batch_size):
        df = pd.read_excel(self.path,sheet_name=None)
        invoice = df[self.name_column_invoice].set_index(self.invoiceFactorNumber)
        invoiceItem = df[self.name_column_invoiceItem].set_index(self.invoiceItemFactorNumber)
        payment = df[self.name_column_payment].set_index(self.PaymentFactorNumber)

        groupItem = invoiceItem.groupby(self.invoiceItemFactorNumber)
        groupPayment = payment.groupby(self.PaymentFactorNumber)

        keys = set(invoice.index).union(set(invoiceItem)).union(set(payment.index))

        for i, key in enumerate(keys):
            if i % batch_size == 0:
                if i > 1:
                    yield batch
                batch = []
            # a = groupItem.get_group()
            data = [invoice.loc[key].to_dict()]
            batch.append(data)
        yield batch

    def readSheet(self,sheet_name):
        try:
            df = pd.read_excel(self.path,sheet_name=sheet_name,dtype=str)
            return df
        except:
            return None

    
    def checkExcellNew(self,type,pattern):
        try:
            # import time
            # start_time = time.time()
            excellFile = pd.ExcelFile(self.path)
            self.sheetNames = excellFile.sheet_names
            # num_threads = len(self.sheetNames)
            self.data = pd.read_excel(self.path,sheet_name=None,dtype=str)
            # with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
            #     self.data = executor.map(self.readSheet,self.sheetNames)
           
            # print("--- %s seconds ---" % (time.time() - start_time))
        
            result = []
            for  index,(sheet_name, df) in enumerate(self.data.items()):
                df_cols = len(df.columns)
                self.cols = InvoiceData(type,pattern,index)
                if df_cols == self.cols.colum:
                    result.append(1)
                else: result.append(0) 
            return result
        except:
            return None
    
    def generate_uuid(self):
        return str(uuid.uuid4())
    
    def convert_date(self,x):
        try:
            date = str(x).split('/')
            return str(jdatetime.date(int(date[0]),int(date[1]),int(date[2])).togregorian())
        except: 
            return None


    def readDataExcel(self,sheet_index,number) -> pd.DataFrame:
        df = self.data[self.sheetNames[sheet_index]]
        df = pd.DataFrame(df)
        df = df.set_axis(col.invoiceItemsGeneral(),axis='columns')
        df = df.replace(np.nan,None)
        df['uniqueId'] = df.apply(lambda x :  self.generate_uuid(),axis=1)
        
        df = df.assign(CooperationCode = "Eitak-" + VersionApp.version)
        b =[]
        for s in df.columns:
            if 'Date' in s or 'date' in s :
                b.append(s)

        df[[s for s in df.columns if 'Date' in s or 'date' in s]] = df[[s for s in df.columns if 'Date' in s or 'date' in s]].applymap(self.convert_date)
        
        df_item = self.data[self.sheetNames[number]]
        df_item = pd.DataFrame(df_item)
        df_item = df_item.replace(np.nan,None)
        
        df_item = df_item.set_axis(col.invoiceItemsGeneral(),axis='columns')

        mearge_data = pd.merge(df, df_item,on=['invoiceNumber','invoiceDate'])

        gp_data = mearge_data.groupby(['invoiceNumber','invoiceDate']).apply(lambda x :{
            **{column : x[column].iloc[0] for column in df.columns},
            'Items' : x[df_item.columns[2:]].to_dict(orient='records') 
        })
        dfJson = gp_data.to_json(orient='records')
        # result = dfJson.t(orient="records")
        print(dfJson)

    def readExcelSheet(self,typeInvoice:int,patternInvoic:int,indexSheet:int,TypeDate:int) -> pd.DataFrame:
        try:
            df = self.data[self.sheetNames[indexSheet]]
            df  =pd.DataFrame(df)
            col = Columns(typeInvoice,patternInvoic,indexSheet)
            df = df.set_axis(col.columnsNames,axis='columns')
            df = df.replace(np.nan,None)
            df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

            if indexSheet == 0:
                df.index = df.index + 2
                df['uniqueId'] = df.apply(lambda x :  self.generate_uuid(),axis=1)
                df = df.assign(CooperationCode = "Eitak-" + VersionApp.version)
                df.insert(loc=0,column='ExcelRowNumber',value=df.index.to_list())

            if TypeDate == 2:
                df[[s for s in df.columns if 'Date' in s or 'date' in s]] = df[[s for s in df.columns if 'Date' in s or 'date' in s]].applymap(self.convert_date)
            
            return df
        except:
            return None

    def PreparationData(self,invoice:pd.DataFrame,item:pd.DataFrame,pay:pd.DataFrame=None) -> pd.DataFrame:
        mearge_data =  pd.merge(invoice,item,on=['invoiceNumber','invoiceDate'])
        if pay is None: 
            gp_data = mearge_data.groupby(['invoiceNumber','invoiceDate']).apply(lambda x :{
            **{column : x[column].iloc[0] for column in invoice.columns[1:]},
            'invoiceItems' : x[item.columns[2:]].to_dict(orient='records') 
                })
            return gp_data
        else:
            group_invoice = pd.merge(mearge_data,pay,on=['invoiceNumber','invoiceDate'])
            gp_data = group_invoice.groupby(['invoiceNumber','invoiceDate']).apply(lambda x :{
            **{column : x[column].iloc[0] for column in invoice.columns[1:]},
            'invoiceItems' : x[item.columns[2:]].to_dict(orient='records'),
            'invoicePayments' : x[pay.columns[2:]].to_dict(orient='records') 
                }).reset_index(drop=True)
            return gp_data

    def invoiceByRowExcell(self, invoice:pd.DataFrame) -> pd.DataFrame:
        invoice.insert(loc=0, column='ExcelRowNumber', value=invoice.index.to_list())
        listSelectName = ['ExcelRowNumber','invoiceNumber','uniqueId']
        for i in invoice.columns.values:
            if not i in listSelectName :
                invoice = invoice.drop(columns=i) 
        return invoice


if __name__ == "__main__":
    ed = ExcellData("./dataTest/Invoice_InvoicePatternId.xlsx")
    ed.checkExcellNew(1,1)
    # pay = ed.readDataExcel(1,1)
    invoice = ed.readExcelSheet(1,1,0,2)
    item = ed.readExcelSheet(1,1,1,2)
    pay = ed.readExcelSheet(1,1,2,2)
    if len(pay) == 0:
        pay = None
    data  = ed.PreparationData(invoice,item,pay)
    dfJson = data.to_json(orient='records')
    print (len(dfJson))


    # for batch_data in ed.data_generator(batch_size=10):
    #     print (batch_data)
    # p = "./data/sampel11.xlsx"
    # p = p[:-5] + "_result.xlsx"  
    # print (p)