import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter

class ExcellData ():
    def __init__(self, path) -> None:
        self.path = path

    
    def getInvoice(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.name_column_invoice, dtype=str)
            data = data.replace(np.nan, '')
            invoice = np.array(data)
            return invoice
        except:
            return []

    def getInvoiceItem(self):
        try:
            data = pd.read_excel(self.path,sheet_name= self.name_column_invoiceItem, dtype=str)
            data = data.replace(np.nan, '')
            item = np.array(data)
            return item
        except:
            return []
    
    def getPayment(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.name_column_payment, dtype=str)
            data = data.replace(np.nan, '')
            pay = np.array(data)
            return pay
        except:
            return []
        
    
        
    def checkExcel(self, paternType:int):
        try:
            data = openpyxl.load_workbook(self.path)
            result = []
            if len(data.sheetnames) > 3 or len(data.sheetnames) < 1:
                return result
            
            index = 0
            for sheet in data.sheetnames:
                columns = data.get_sheet_by_name(sheet)
                if index == 0 and paternType == 1:
                    if columns.max_column == 21 : 
                        result.append(1) 
                        self.name_column_invoice = sheet
                    else: result.append(0) 
                elif index == 0 and paternType == 2:
                    if columns.max_column == 16 : 
                        result.append(1) 
                        self.name_column_invoice = sheet
                    else: result.append(0) 
                elif index == 1 :
                    if columns.max_column == 20 : 
                        result.append(1) 
                        self.name_column_invoiceItem = sheet
                    else: result.append(0) 
                elif index == 2 : 
                    if columns.max_column == 11 : 
                        result.append(1) 
                        self.name_column_payment = sheet
                    else: result.append(0) 
                
                index += 1
            return result
        except:
            return None
        # if len(data.sheetnames) == 3:
        #     self.name_column_invoice = data.sheetnames[0]
        #     self.name_column_invoiceItem = data.sheetnames[1]
        #     self.name_column_payment = data.sheetnames[2]
        #     if paternType == 1:
        #         c = data.get_sheet_by_name(self.name_column_invoice)
        #         print(c)  
        # elif len(data.sheetnames) == 2:
        #     self.name_column_invoice = data.sheetnames[0]
        #     self.name_column_invoiceItem = data.sheetnames[1]
        #     if paternType == 1:
        #         c = data.get_sheet_by_name(self.name_column_invoice)
        #         print(c.max_column)  
        # else:
        #     return False
        # data._sheets[0]._parent.worksheets[0].max_column
    @classmethod
    def getLetter(col):
        return get_column_letter(col)
    
    def preparationExcellN11(self):

        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(22)+'1'] = "uniqueId"
        worksheet[get_column_letter(23)+'1'] = "Status Request Server"
        worksheet[get_column_letter(24)+'1'] = "taxSerialNumber"
        worksheet[get_column_letter(25)+'1'] = "message"
        worksheet[get_column_letter(26)+'1'] = "FieldError"
        workbook.save(self.path)



    def preparationExcellN21(self):
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(17)+'1'] = "uniqueId"
        worksheet[get_column_letter(18)+'1'] = "Status Request Server"
        worksheet[get_column_letter(19)+'1'] = "taxSerialNumber"
        worksheet[get_column_letter(20)+'1'] = "message"
        worksheet[get_column_letter(21)+'1'] = "FieldError"
        workbook.save(self.path)
        
    def SaveResultN11(self,data,index):
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(22)+str(index+1)] = data[0]#"uniqueId"
        worksheet[get_column_letter(23)+str(index+1)] = data[1]#"Status Request Server"
        worksheet[get_column_letter(24)+str(index+1)] = data[2]#"taxSerialNumber"
        worksheet[get_column_letter(25)+str(index+1)] = data[3]#"message"
        worksheet[get_column_letter(26)+str(index+1)] = data[4]#"FieldError"
        workbook.save(self.path)

    def SaveResultN21(self,data,index):
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(17)+str(index+1)] = data[0]#"uniqueId"
        worksheet[get_column_letter(18)+str(index+1)] = data[1]#"Status Request Server"
        worksheet[get_column_letter(19)+str(index+1)] = data[2]#"taxSerialNumber"
        worksheet[get_column_letter(20)+str(index+1)] = data[3]#"message"
        worksheet[get_column_letter(21)+str(index+1)] = data[4]#"FieldError
        workbook.save(self.path)
    

if __name__ == "__main__":
    ed = ExcellData("./data/sampel11.xl")
    a = ed.checkExcel(2)
    print (a)