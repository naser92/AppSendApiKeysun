import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter
import shutil
import os

class ExcellData ():
    def __init__(self, path) -> None:
        self.path = path
          
    
    def getInvoice(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.name_column_invoice, dtype=str)
            data = data.replace(np.nan, '')
            invoice = np.array(data)
            return invoice
        except:
            return []

    def getInvoiceItem(self):
        try:
            data = pd.read_excel(self.path,sheet_name= self.name_column_invoiceItem, dtype=str)
            data = data.replace(np.nan, '')
            item = np.array(data)
            return item
        except:
            return []
    
    def getPayment(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.name_column_payment, dtype=str)
            data = data.replace(np.nan, '')
            pay = np.array(data)
            return pay
        except:
            return []
        
    def getRevokeInvoice(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.name_column_revoke, dtype=str)
            data = data.replace(np.nan, '')
            pay = np.array(data)
            return pay
        except:
            return []
        
    def getDeleteUniquId(self):
        try:
            data = pd.read_excel(self.path,sheet_name=self.name_column_delete, dtype=str)
            data = data.replace(np.nan, '')
            pay = np.array(data)
            return pay
        except:
            return []
    
        
    def checkExcel(self, paternType:int):
        try:
            data = openpyxl.load_workbook(self.path)
            result = []
            if len(data.sheetnames) > 3 or len(data.sheetnames) < 1:
                return result
            
            index = 0
            for sheet in data.sheetnames:
                columns = data.get_sheet_by_name(sheet)
                if index == 0 and paternType == 1:
                    if columns.max_column == 21 : 
                        result.append(1) 
                        self.name_column_invoice = sheet
                    else: result.append(0) 
                elif index == 0 and paternType == 2:
                    if columns.max_column == 16 : 
                        result.append(1) 
                        self.name_column_invoice = sheet
                    else: result.append(0) 
                elif index == 1 :
                    if columns.max_column == 20 : 
                        result.append(1) 
                        self.name_column_invoiceItem = sheet
                    else: result.append(0) 
                elif index == 2 : 
                    if columns.max_column == 11 : 
                        result.append(1) 
                        self.name_column_payment = sheet
                    else: result.append(0) 
                
                index += 1
            return result
        except:
            return None
        
    def excelCheckRevok(self):
        try:
            data = openpyxl.load_workbook(self.path)
            flag = False
            if len(data.sheetnames) > 1:
                return flag
            sheet = data.sheetnames[0]
            columns = data.get_sheet_by_name(sheet)
            if columns.max_column == 7:
                self.name_column_revoke = sheet
                flag = True
            
            return flag

        except:
            return False
        
    def excelCheckDelete(self):
        try:
            data = openpyxl.load_workbook(self.path)
            flag = False
            sheet = data.sheetnames[0]
            columns = data.get_sheet_by_name(sheet)
            if columns.max_column == 1:
                self.name_column_delete = sheet
                flag = True
            return flag
        except:
            return False
    
    @classmethod
    def getLetter(col):
        return get_column_letter(col)
    
    def copyFile(self,path):
        try:
            new  = path[:-5] + "_read.xlsx"
            result = path[:-5] + "_result.xlsx"
            shutil.copyfile(path, new,follow_symlinks=True)
            shutil.copyfile(path, result,follow_symlinks=True)
            os.remove(path)
            self.path = result
        except:
            self.path = path
    
    def preparationExcellN11(self):
        self.copyFile(self.path)
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(22)+'1'] = "uniqueId"
        worksheet[get_column_letter(23)+'1'] = "Status Request Server"
        worksheet[get_column_letter(24)+'1'] = "taxSerialNumber"
        worksheet[get_column_letter(25)+'1'] = "message"
        worksheet[get_column_letter(26)+'1'] = "FieldError"
        workbook.save(self.path)



    def preparationExcellN21(self):
        self.copyFile(self.path)
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(17)+'1'] = "uniqueId"
        worksheet[get_column_letter(18)+'1'] = "Status Request Server"
        worksheet[get_column_letter(19)+'1'] = "taxSerialNumber"
        worksheet[get_column_letter(20)+'1'] = "message"
        worksheet[get_column_letter(21)+'1'] = "FieldError"
        workbook.save(self.path)
        
    def SaveResultN11(self,data,index):
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(22)+str(index+1)] = data[0]#"uniqueId"
        worksheet[get_column_letter(23)+str(index+1)] = data[1]#"Status Request Server"
        worksheet[get_column_letter(24)+str(index+1)] = data[2]#"taxSerialNumber"
        worksheet[get_column_letter(25)+str(index+1)] = data[3]#"message"
        worksheet[get_column_letter(26)+str(index+1)] = data[4]#"FieldError"
        workbook.save(self.path)

    def SaveResultN21(self,data,index):
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[self.name_column_invoice]
        worksheet[get_column_letter(17)+str(index+1)] = data[0]#"uniqueId"
        worksheet[get_column_letter(18)+str(index+1)] = data[1]#"Status Request Server"
        worksheet[get_column_letter(19)+str(index+1)] = data[2]#"taxSerialNumber"
        worksheet[get_column_letter(20)+str(index+1)] = data[3]#"message"
        worksheet[get_column_letter(21)+str(index+1)] = data[4]#"FieldError
        workbook.save(self.path)
    

if __name__ == "__main__":
    ed = ExcellData("./data/sampel11.xlsx")
    # a = ed.checkExcel(2)
    p = "./data/sampel11.xlsx"
    p = p[:-5] + "_result.xlsx"  
    print (p)